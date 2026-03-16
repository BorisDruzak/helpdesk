"""
Модуль экспорта документов в Excel/Calc
Поддерживает выбор периода, полей и экспорт базы данных
"""

import os
import shutil
from datetime import datetime, date
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QListWidget, QListWidgetItem, QMessageBox, QFileDialog,
    QCheckBox, QGroupBox, QProgressDialog, QTabWidget, QWidget,
    QDateEdit, QRadioButton, QButtonGroup
)
from PyQt5.QtCore import Qt, pyqtSignal, QDate
from PyQt5.QtGui import QFont as QFont_GUI
import tarfile  # для создания .tar.gz архивов
import platform  # для определения ОС
from ui_styles import AppStyles
class ExportDialog(QDialog):
    """Главный диалог экспорта с вкладками"""
    backup_completed = pyqtSignal()
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.preset_documents = []  # предзаполненные документы
        self.preset_mode = None     # режим ('all_from_table' или 'selected')
        self.setWindowTitle("📤 Экспорт данных")
        self.setMinimumSize(600, 700)
        self.init_ui()
        self.backup_tab = self.create_backup_tab()
        self.tab_widget.addTab(self.backup_tab, "📦 Архивация")
    def init_ui(self):
        """Инициализация интерфейса"""
        layout = QVBoxLayout()
        
        # Создаем вкладки
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet(AppStyles.tab_widget())
        
        # Вкладка 1: Экспорт в печатную форму
        self.print_form_tab = self.create_print_form_tab()
        self.tab_widget.addTab(self.print_form_tab, "📄 Печатная форма")
        
        
        
        layout.addWidget(self.tab_widget)
        
        # Кнопки диалога
        buttons_layout = QHBoxLayout()
        
        export_btn = QPushButton("✅ Экспортировать")
        export_btn.setStyleSheet(AppStyles.button_success())
        export_btn.clicked.connect(self.do_export)
        
        cancel_btn = QPushButton("❌ Отмена")
        cancel_btn.setStyleSheet(AppStyles.button_danger())
        cancel_btn.clicked.connect(self.reject)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(export_btn)
        buttons_layout.addWidget(cancel_btn)
        
        layout.addLayout(buttons_layout)
        
        self.setLayout(layout)
    def create_backup_tab(self) -> QWidget:
        """Создать вкладку архивации базы данных"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Заголовок
        title = QLabel("📦 Архивация базы данных")
        title.setFont(QFont_GUI("Segoe UI", 12, QFont_GUI.Bold))
        title.setStyleSheet("color: #2c3e50; padding: 10px;")
        layout.addWidget(title)
        
        # Описание
        desc_text = """
        <p style='color: #555; font-size: 10pt;'>
        Создание полного архива базы данных включая:<br>
        • Файл базы данных (.db)<br>
        • Все прикрепленные документы (опционально)<br>
        • Исполняемый файл программы (опционально)<br>
        <br>
        Формат архива: <b>.tar.gz</b> (универсальный для Windows и Linux)<br>
        Название: <b>(имя_бд)_(дата).tar.gz</b><br>
        </p>
        """
        desc = QLabel(desc_text)
        desc.setWordWrap(True)
        desc.setStyleSheet(AppStyles.lable())
        layout.addWidget(desc)
        
        # Информация о последней архивации
        backup_info_group = QGroupBox("📊 Информация об архивации")
        backup_info_group.setStyleSheet(AppStyles.group_box())
        backup_info_layout = QVBoxLayout()
        
        # Получаем статистику
        backup_stats = self.db_manager.get_backup_statistics()
        last_backup = backup_stats.get('last_backup')
        backup_count = backup_stats.get('backup_count', 0)
        days_since = backup_stats.get('days_since_backup')
        
        if last_backup:
            try:
                backup_date = datetime.strptime(last_backup, "%Y-%m-%d %H:%M:%S")
                formatted_date = backup_date.strftime("%d.%m.%Y %H:%M")
            except:
                formatted_date = last_backup
            
            # Определяем цвет предупреждения
            if days_since is not None:
                if days_since >= 7:
                    warning_color = "#f44336"  # Красный - больше недели
                    warning_text = f"⚠️ Прошло {days_since} дней с последней архивации!"
                elif days_since >= 5:
                    warning_color = "#ff9800"  # Оранжевый - скоро неделя
                    warning_text = f"⏰ Прошло {days_since} дней с последней архивации"
                else:
                    warning_color = "#4caf50"  # Зеленый - все ок
                    warning_text = f"✅ Прошло {days_since} дней с последней архивации"
            else:
                warning_color = "#666"
                warning_text = ""
            
            info_text = f"""
            <table style='width: 100%; font-size: 10pt;'>
            <tr><td><b>Последняя архивация:</b></td><td>{formatted_date}</td></tr>
            <tr><td><b>Всего архивов создано:</b></td><td>{backup_count}</td></tr>
            </table>
            <p style='color: {warning_color}; font-weight: bold; margin-top: 10px;'>
            {warning_text}
            </p>
            """
        else:
            info_text = """
            <p style='color: #f44336; font-weight: bold; font-size: 11pt;'>
            ⚠️ Архивация еще не выполнялась!<br>
            Рекомендуется создать резервную копию данных.
            </p>
            """
        
        info_label = QLabel(info_text)
        info_label.setStyleSheet(AppStyles.lable())
        info_label.setTextFormat(Qt.RichText)
        backup_info_layout.addWidget(info_label)
        
        backup_info_group.setLayout(backup_info_layout)
        layout.addWidget(backup_info_group)
        
        # Информация о текущей БД
        db_info_group = QGroupBox("📂 Данные для архивации")
        db_info_group.setStyleSheet(AppStyles.group_box())
        db_info_layout = QVBoxLayout()
        
        stats = self.db_manager.get_documents_statistics()
        total_docs = stats.get('total_documents', 0)
        
        db_path = self.db_manager.db_path
        db_name = os.path.basename(os.path.dirname(db_path))
        db_size = os.path.getsize(db_path) if os.path.exists(db_path) else 0
        db_size_mb = db_size / (1024 * 1024)
        
        # ✅ ОПТИМИЗАЦИЯ: НЕ считаем размер сразу (экономим 20-40 секунд)
        files_path = self.db_manager.files_base_path
        files_exists = os.path.exists(files_path)

        # Показываем информацию без подсчета размера
        db_info_text = f"""
        <table style='width: 100%; font-size: 10pt;'>
        <tr><td><b>Имя БД:</b></td><td>{db_name}</td></tr>
        <tr><td><b>Размер БД:</b></td><td>{db_size_mb:.2f} МБ</td></tr>
        <tr><td><b>Документов:</b></td><td>{total_docs}</td></tr>
        <tr><td><b>Папка файлов:</b></td><td>{'✅ Существует' if files_exists else '❌ Не найдена'}</td></tr>
        <tr><td colspan='2' style='padding-top: 10px;'><i>💡 Размер файлов будет подсчитан при архивации</i></td></tr>
        </table>
        """
        
        db_info_label = QLabel(db_info_text)
        db_info_label.setStyleSheet(AppStyles.lable())
        db_info_label.setTextFormat(Qt.RichText)
        db_info_layout.addWidget(db_info_label)
        
        db_info_group.setLayout(db_info_layout)
        layout.addWidget(db_info_group)
        
        # Опции архивации
        options_group = QGroupBox("⚙️ Настройки архивации")
        options_group.setStyleSheet(AppStyles.group_box())
        options_layout = QVBoxLayout()
        
        self.backup_include_files = QCheckBox("📁 Включить все прикрепленные документы")
        self.backup_include_files.setChecked(True)
        self.backup_include_files.setStyleSheet(AppStyles.check())
        options_layout.addWidget(self.backup_include_files)
        
        self.backup_include_exe = QCheckBox("💻 Включить исполняемый файл программы")
        self.backup_include_exe.setChecked(False)
        self.backup_include_exe.setStyleSheet(AppStyles.check())
        options_layout.addWidget(self.backup_include_exe)
        
        # Информация об exe
        exe_info = QLabel("   ℹ️ Будет включен файл из папки приложения")
        exe_info.setStyleSheet(AppStyles.lable())
        options_layout.addWidget(exe_info)
        
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        layout.addStretch()
        widget.setLayout(layout)
        
        return widget
    def create_print_form_tab(self) -> QWidget:
        """Создать вкладку экспорта в печатную форму"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Заголовок
        title = QLabel("📄 Экспорт документов в печатную форму")
        title.setFont(QFont_GUI("Segoe UI", 12, QFont_GUI.Bold))
        title.setStyleSheet(AppStyles.lable())
        layout.addWidget(title)
        
        # === БЛОК ВЫБОРА ПЕРИОДА ===
        period_group = QGroupBox("📅 Выберите период экспорта")
        period_group.setStyleSheet(AppStyles.group_box())
        period_layout = QVBoxLayout()
        
        # Радио-кнопки выбора режима
        self.period_button_group = QButtonGroup()
        
        # Режим: один день
        self.single_day_radio = QRadioButton("📆 Один день")
        self.single_day_radio.setChecked(True)
        self.single_day_radio.toggled.connect(self.on_period_mode_changed)
        self.period_button_group.addButton(self.single_day_radio)
        period_layout.addWidget(self.single_day_radio)
        
        # Выбор одного дня
        single_day_layout = QHBoxLayout()
        single_day_layout.addSpacing(30)
        single_day_label = QLabel("Дата:")
        self.single_date_edit = QDateEdit()
        self.single_date_edit.setCalendarPopup(True)
        self.single_date_edit.setDate(QDate.currentDate())
        self.single_date_edit.setDisplayFormat("dd.MM.yyyy")
        self.single_date_edit.setStyleSheet(AppStyles.lable())
        single_day_layout.addWidget(single_day_label)
        single_day_layout.addWidget(self.single_date_edit)
        single_day_layout.addStretch()
        period_layout.addLayout(single_day_layout)
        
        # Режим: диапазон дат
        self.date_range_radio = QRadioButton("📊 Диапазон дат")
        self.period_button_group.addButton(self.date_range_radio)
        period_layout.addWidget(self.date_range_radio)
        
        # Выбор диапазона
        range_layout = QHBoxLayout()
        range_layout.addSpacing(30)
        
        from_label = QLabel("С:")
        self.date_from_edit = QDateEdit()
        self.date_from_edit.setCalendarPopup(True)
        self.date_from_edit.setDate(QDate.currentDate().addMonths(-1))
        self.date_from_edit.setDisplayFormat("dd.MM.yyyy")
        self.date_from_edit.setEnabled(False)
        self.date_from_edit.setStyleSheet(AppStyles.lable())
        
        to_label = QLabel("По:")
        self.date_to_edit = QDateEdit()
        self.date_to_edit.setCalendarPopup(True)
        self.date_to_edit.setDate(QDate.currentDate())
        self.date_to_edit.setDisplayFormat("dd.MM.yyyy")
        self.date_to_edit.setEnabled(False)
        self.date_to_edit.setStyleSheet(AppStyles.lable())
        
        range_layout.addWidget(from_label)
        range_layout.addWidget(self.date_from_edit)
        range_layout.addWidget(to_label)
        range_layout.addWidget(self.date_to_edit)
        range_layout.addStretch()
        period_layout.addLayout(range_layout)
        
        # Режим: все документы
        self.all_docs_radio = QRadioButton("📚 Все документы в базе")
        self.period_button_group.addButton(self.all_docs_radio)
        period_layout.addWidget(self.all_docs_radio)
        
        # Режим предустановленных документов (из таблицы или выделенных)
        self.preset_docs_radio = QRadioButton("📋 Документы из таблицы")
        self.period_button_group.addButton(self.preset_docs_radio)
        period_layout.addWidget(self.preset_docs_radio)
        
        # Информация о предустановленных документах
        self.preset_info_label = QLabel("")
        self.preset_info_label.setStyleSheet(AppStyles.lable())
        self.preset_info_label.setVisible(False)
        period_layout.addWidget(self.preset_info_label)
        
        period_group.setLayout(period_layout)
        layout.addWidget(period_group)
        
        # === БЛОК ВЫБОРА ПОЛЕЙ ===
        fields_group = QGroupBox("📋 Выберите поля для экспорта")
        fields_group.setStyleSheet(AppStyles.group_box())
        fields_layout = QVBoxLayout()
        
        # Описание
        desc = QLabel("Отметьте галочками поля для включения в таблицу")
        desc.setStyleSheet("color: #666; font-size: 9pt;")
        fields_layout.addWidget(desc)
        
        # Список полей
        self.fields_list = self._create_fields_list()
        fields_layout.addWidget(self.fields_list)
        
        # Кнопки управления выбором
        fields_buttons_layout = QHBoxLayout()
        
        select_all_btn = QPushButton("✅ Все")
        select_all_btn.setStyleSheet(AppStyles.butt())
        select_all_btn.clicked.connect(self.select_all_fields)
        
        deselect_all_btn = QPushButton("❌ Снять")
        deselect_all_btn.setStyleSheet(AppStyles.butt())
        deselect_all_btn.clicked.connect(self.deselect_all_fields)
        
        default_btn = QPushButton("🔄 По умолчанию")
        default_btn.setStyleSheet(AppStyles.butt())
        default_btn.clicked.connect(self.select_default_fields)
        
        fields_buttons_layout.addWidget(select_all_btn)
        fields_buttons_layout.addWidget(deselect_all_btn)
        fields_buttons_layout.addWidget(default_btn)
        fields_buttons_layout.addStretch()
        
        fields_layout.addLayout(fields_buttons_layout)
        
        fields_group.setLayout(fields_layout)
        layout.addWidget(fields_group)
        
        layout.addStretch()
        widget.setLayout(layout)
        
        # Устанавливаем поля по умолчанию
        self.select_default_fields()
        
        return widget
    
    def create_db_export_tab(self) -> QWidget:
        """Создать вкладку экспорта базы данных"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Заголовок
        title = QLabel("💾 Экспорт базы данных")
        title.setFont(QFont_GUI("Segoe UI", 12, QFont_GUI.Bold))
        title.setStyleSheet("color: #2c3e50; padding: 10px;")
        layout.addWidget(title)
        
        # Описание
        desc_text = """
        <p style='color: #555; font-size: 10pt;'>
        Экспорт создаст полную копию базы данных включая:<br>
        • Файл базы данных (.db)<br>
        • Все прикрепленные документы<br>
        • Структуру папок по годам и месяцам<br>
        <br>
        Эту копию можно использовать для:<br>
        ✓ Резервного копирования<br>
        ✓ Переноса на другой компьютер<br>
        ✓ Архивирования данных<br>
        </p>
        """
        desc = QLabel(desc_text)
        desc.setWordWrap(True)
        desc.setStyleSheet(AppStyles.lable())
        layout.addWidget(desc)
        
        # Информация о текущей БД
        info_group = QGroupBox("📊 Информация о базе данных")
        info_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
        """)
        info_layout = QVBoxLayout()
        
        # Получаем статистику
        stats = self.db_manager.get_documents_statistics()
        total_docs = stats.get('total_documents', 0)
        
        db_path = self.db_manager.db_path
        db_size = os.path.getsize(db_path) if os.path.exists(db_path) else 0
        db_size_mb = db_size / (1024 * 1024)
        
        info_text = f"""
        <table style='width: 100%; font-size: 10pt;'>
        <tr><td><b>Путь к БД:</b></td><td>{db_path}</td></tr>
        <tr><td><b>Размер БД:</b></td><td>{db_size_mb:.2f} МБ</td></tr>
        <tr><td><b>Документов:</b></td><td>{total_docs}</td></tr>
        <tr><td><b>Папка файлов:</b></td><td>{self.db_manager.files_base_path}</td></tr>
        </table>
        """
        
        info_label = QLabel(info_text)
        info_label.setTextFormat(Qt.RichText)
        info_layout.addWidget(info_label)
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # Опции экспорта
        options_group = QGroupBox("⚙️ Настройки экспорта")
        options_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
        """)
        options_layout = QVBoxLayout()
        
        self.include_files_checkbox = QCheckBox("📁 Включить прикрепленные документы")
        self.include_files_checkbox.setChecked(True)
        self.include_files_checkbox.setStyleSheet("font-size: 10pt;")
        options_layout.addWidget(self.include_files_checkbox)
        
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        layout.addStretch()
        widget.setLayout(layout)
        
        return widget
    
    def on_period_mode_changed(self):
        """Обработчик изменения режима выбора периода"""
        is_single_day = self.single_day_radio.isChecked()
        is_date_range = self.date_range_radio.isChecked()
        is_preset = self.preset_docs_radio.isChecked()  # ← НОВОЕ
        
        self.single_date_edit.setEnabled(is_single_day)
        self.date_from_edit.setEnabled(is_date_range)
        self.date_to_edit.setEnabled(is_date_range)
        
        # Показываем информацию о предустановленных документах
        if is_preset:
            count = len(self.preset_documents)
            
            # Определяем текст в зависимости от режима
            if self.preset_mode == 'selected':
                mode_text = "выделенных"
                icon = "✅"
            else:
                mode_text = "из таблицы"
                icon = "📋"
            
            self.preset_info_label.setText(f"   {icon} Документов {mode_text}: {count}")
            self.preset_info_label.setVisible(True)
        else:
            self.preset_info_label.setVisible(False)
    
    def _create_fields_list(self) -> QListWidget:
        """Создать список полей"""
        list_widget = QListWidget()
        list_widget.setStyleSheet(AppStyles.list_widget())
        
        # Все доступные поля
        self.available_fields = {
            'reg_number': 'Рег. номер',
            'reg_date': 'Рег. дата',
            'title': 'Заголовок',
            'number': 'Номер',
            'status_name': 'Статус',
            'type_name': 'Тип документа',
            'document_kind_name': 'Вид документа',
            'signing_type_name': 'Тип подписания',
            'executor_name': 'Исполнитель',
            'responsible_executor_name': 'Ответственный исполнитель',
            'theme_name': 'Тема',
            'signers': 'Кто подписал',
            'approvers': 'Кто согласовал',
            'should_publish': 'Подлежит опубликованию',
            'published_where_name': 'Где опубликовано',
            'published_date': 'Дата публикации',
            'control_date': 'Дата контроля',
            'removed_from_control': 'Снято с контроля',
            'execution_result': 'Результат исполнения',
            'pages_count': 'Количество страниц',
            'attachments_count': 'Количество приложений',
            'case_number': 'Номер дела',
            'volume_number': 'Номер тома',
            'sheets': 'Листы'
        }
        
        # Поля по умолчанию
        self.default_fields = [
            'reg_number', 'reg_date', 'title', 'signers', 
            'executor_name', 'control_date'
        ]
        
        for field_key, field_name in self.available_fields.items():
            item = QListWidgetItem(field_name)
            item.setData(Qt.UserRole, field_key)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            list_widget.addItem(item)
        
        return list_widget
    
    def select_all_fields(self):
        """Выбрать все поля"""
        for i in range(self.fields_list.count()):
            item = self.fields_list.item(i)
            item.setCheckState(Qt.Checked)
    
    def deselect_all_fields(self):
        """Снять выбор со всех полей"""
        for i in range(self.fields_list.count()):
            item = self.fields_list.item(i)
            item.setCheckState(Qt.Unchecked)
    
    def select_default_fields(self):
        """Выбрать поля по умолчанию"""
        self.deselect_all_fields()
        for i in range(self.fields_list.count()):
            item = self.fields_list.item(i)
            field_key = item.data(Qt.UserRole)
            if field_key in self.default_fields:
                item.setCheckState(Qt.Checked)
    
    def get_selected_fields(self) -> List[tuple]:
        """Получить выбранные поля"""
        selected = []
        for i in range(self.fields_list.count()):
            item = self.fields_list.item(i)
            if item.checkState() == Qt.Checked:
                field_key = item.data(Qt.UserRole)
                field_name = item.text()
                selected.append((field_key, field_name))
        return selected
    
    def get_documents_for_period(self) -> List[Dict]:
        """Получить документы за выбранный период"""
        try:
            if self.preset_docs_radio.isChecked():
                if not self.preset_documents:
                    QMessageBox.warning(
                        self,
                        "⚠️ Нет документов",
                        "Документы для экспорта не установлены"
                    )
                    return []
                
                print(f"✅ Используется {len(self.preset_documents)} предустановленных документов")
                return self.preset_documents
            # Определяем режим выбора периода
            if self.single_day_radio.isChecked():
                # Один день
                selected_date = self.single_date_edit.date().toString("yyyy-MM-dd")
                query = """
                    SELECT d.* FROM documents d
                    WHERE d.reg_date = ?
                    ORDER BY d.reg_date, d.id
                """
                params = (selected_date,)
                
            elif self.date_range_radio.isChecked():
                # Диапазон дат
                date_from = self.date_from_edit.date().toString("yyyy-MM-dd")
                date_to = self.date_to_edit.date().toString("yyyy-MM-dd")
                
                if date_from > date_to:
                    QMessageBox.warning(
                        self,
                        "⚠️ Ошибка",
                        "Дата начала не может быть позже даты окончания"
                    )
                    return []
                
                query = """
                    SELECT d.* FROM documents d
                    WHERE d.reg_date BETWEEN ? AND ?
                    ORDER BY d.reg_date, d.id
                """
                params = (date_from, date_to)
                
            else:
                # Все документы
                query = """
                    SELECT d.* FROM documents d
                    ORDER BY d.reg_date DESC, d.id DESC
                """
                params = ()
            
            # Выполняем запрос
            cursor = self.db_manager.connection.cursor()
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # Преобразуем в список словарей и дополняем данными
            documents = []
            for row in rows:
                doc_dict = dict(row)
                doc_id = doc_dict['id']
                
                # Дополняем полными данными
                full_doc = self.db_manager.get_document_by_id(doc_id)
                if full_doc:
                    documents.append(full_doc)
            
            return documents
            
        except Exception as e:
            print(f"❌ Ошибка получения документов: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def do_export(self):
        """Выполнить экспорт"""
        current_tab = self.tab_widget.currentIndex()
        
        if current_tab == 0:
            # Экспорт в печатную форму
            self.export_print_form()
        else:
            # Экспорт базы данных
            self.create_backup()
    def create_backup(self):
        """Создать архив базы данных"""
        try:
            # Генерируем имя архива
            db_name = os.path.basename(os.path.dirname(self.db_manager.db_path))
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            archive_name = f"{db_name}_{timestamp}.tar.gz"
            
            # Выбор места сохранения
            save_folder = QFileDialog.getExistingDirectory(
                self,
                "Выберите папку для сохранения архива",
                os.path.dirname(self.db_manager.db_path),
                QFileDialog.ShowDirsOnly
            )
            
            if not save_folder:
                return  # Пользователь отменил
            
            archive_path = os.path.join(save_folder, archive_name)
            
            # Подтверждение
            reply = QMessageBox.question(
                self,
                "📦 Подтверждение архивации",
                f"Создать архив базы данных?\n\n"
                f"Файл: {archive_name}\n"
                f"Путь: {save_folder}\n\n"
                f"Включить документы: {'Да' if self.backup_include_files.isChecked() else 'Нет'}\n"
                f"Включить exe: {'Да' if self.backup_include_exe.isChecked() else 'Нет'}",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply != QMessageBox.Yes:
                return
            
            # Прогресс-бар
            progress = QProgressDialog(
                "Создание архива...",
                "Отмена",
                0, 100,
                self
            )
            progress.setWindowModality(Qt.WindowModal)
            progress.setWindowTitle("📦 Архивация")
            progress.setMinimumDuration(0)
            progress.show()
            
            # Создаем архив
            progress.setValue(5)
            progress.setLabelText("Подготовка файлов...")
            
            # Список файлов для архивации
            files_to_archive = []
            base_dir = os.path.dirname(self.db_manager.db_path)
            
            # 1. Файл БД (обязательно)
            db_file = self.db_manager.db_path
            if os.path.exists(db_file):
                files_to_archive.append((db_file, os.path.basename(db_file)))
            
            # 2. Файлы документов (если выбрано)
            if self.backup_include_files.isChecked():
                progress.setValue(10)
                progress.setLabelText("Добавление файлов документов...")
                
                files_path = self.db_manager.files_base_path
                if os.path.exists(files_path):
                    for root, dirs, files in os.walk(files_path):
                        for file in files:
                            full_path = os.path.join(root, file)
                            # Относительный путь от base_dir
                            arcname = os.path.relpath(full_path, base_dir)
                            files_to_archive.append((full_path, arcname))
            
            # 3. Exe файл (если выбрано)
            if self.backup_include_exe.isChecked():
                progress.setValue(15)
                progress.setLabelText("Добавление исполняемого файла...")
                
                # Ищем exe в папке приложения
                app_dir = os.path.dirname(os.path.dirname(base_dir))
                exe_candidates = [
                    os.path.join(app_dir, "acti_v2.exe"),
                    os.path.join(app_dir, "main.exe"),
                    os.path.join(app_dir, "app.exe")
                ]
                
                for exe_path in exe_candidates:
                    if os.path.exists(exe_path):
                        files_to_archive.append((exe_path, os.path.basename(exe_path)))
                        break
            
            # Создаем tar.gz архив
            progress.setValue(20)
            progress.setLabelText("Создание архива...")
            
            total_files = len(files_to_archive)
            
            with tarfile.open(archive_path, "w:gz") as tar:
                for idx, (filepath, arcname) in enumerate(files_to_archive):
                    if progress.wasCanceled():
                        # Удаляем неполный архив
                        if os.path.exists(archive_path):
                            os.remove(archive_path)
                        return
                    
                    tar.add(filepath, arcname=arcname)
                    
                    # Обновляем прогресс
                    progress_value = 20 + int((idx / total_files) * 70)
                    progress.setValue(progress_value)
                    progress.setLabelText(f"Архивирование: {idx + 1}/{total_files}")
            
            # Обновляем дату архивации в БД
            progress.setValue(95)
            progress.setLabelText("Обновление даты архивации...")
            
            if self.db_manager.update_backup_date():
                print("✅ Дата архивации обновлена в БД")
            else:
                print("⚠️ Не удалось обновить дату архивации")
            
            progress.setValue(100)
            progress.close()
            self.backup_completed.emit()
            # Получаем размер архива
            archive_size = os.path.getsize(archive_path) / (1024 * 1024)
            
            # Сообщение об успехе
            QMessageBox.information(
                self,
                "✅ Архивация завершена",
                f"Архив успешно создан:\n\n"
                f"📁 Файл: {archive_name}\n"
                f"📊 Размер: {archive_size:.2f} МБ\n"
                f"📂 Папка: {save_folder}\n\n"
                f"Файлов в архиве: {total_files}"
            )
            
            self.accept()
            
            # Открыть папку
            reply = QMessageBox.question(
                self,
                "📂 Открыть папку?",
                "Открыть папку с архивом?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.open_folder(save_folder)
        
        except Exception as e:
            if 'progress' in locals():
                progress.close()
            
            QMessageBox.critical(
                self,
                "❌ Ошибка архивации",
                f"Не удалось создать архив:\n\n{str(e)}"
            )
            print(f"❌ Ошибка архивации: {e}")
            import traceback
            traceback.print_exc()
    def export_print_form(self):
        """Экспорт в печатную форму"""
        try:
            # Проверяем выбор полей
            selected_fields = self.get_selected_fields()
            if not selected_fields:
                QMessageBox.warning(
                    self,
                    "⚠️ Предупреждение",
                    "Не выбрано ни одного поля для экспорта"
                )
                return
            
            # Получаем документы за период
            documents = self.get_documents_for_period()
            
            if not documents:
                QMessageBox.warning(
                    self,
                    "⚠️ Нет данных",
                    "За выбранный период документов не найдено"
                )
                return
            
            # Создаем менеджер экспорта
            export_manager = ExportManager(self.db_manager)
            
            # Выполняем экспорт
            filepath = export_manager.export_documents(
                documents, 
                selected_fields, 
                self
            )
            
            if filepath:
                self.accept()
                
                # Предлагаем открыть файл
                reply = QMessageBox.question(
                    self,
                    "📂 Открыть файл?",
                    "Реестр создан успешно!\n\nОткрыть файл в Excel/Calc?",
                    QMessageBox.Yes | QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    self.open_file(filepath)
        
        except Exception as e:
            QMessageBox.critical(
                self,
                "❌ Ошибка",
                f"Не удалось выполнить экспорт:\n\n{str(e)}"
            )
            print(f"❌ Ошибка экспорта: {e}")
            import traceback
            traceback.print_exc()
    
    def export_database(self):
        """Экспорт базы данных"""
        try:
            # Выбор папки для сохранения
            default_name = f"DB_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            save_folder = QFileDialog.getExistingDirectory(
                self,
                "Выберите папку для экспорта базы данных",
                os.path.dirname(self.db_manager.db_path),
                QFileDialog.ShowDirsOnly
            )
            
            if not save_folder:
                return  # Пользователь отменил
            
            # Создаем папку для экспорта
            export_folder = os.path.join(save_folder, default_name)
            os.makedirs(export_folder, exist_ok=True)
            
            # Прогресс
            progress = QProgressDialog(
                "Экспорт базы данных...",
                "Отмена",
                0, 100,
                self
            )
            progress.setWindowModality(Qt.WindowModal)
            progress.setWindowTitle("💾 Экспорт")
            progress.show()
            
            # Копируем файл БД
            progress.setValue(10)
            progress.setLabelText("Копирование файла базы данных...")
            
            db_filename = os.path.basename(self.db_manager.db_path)
            dest_db_path = os.path.join(export_folder, db_filename)
            shutil.copy2(self.db_manager.db_path, dest_db_path)
            
            # Копируем файлы документов если нужно
            if self.include_files_checkbox.isChecked():
                progress.setValue(30)
                progress.setLabelText("Копирование файлов документов...")
                
                files_source = self.db_manager.files_base_path
                files_dest = os.path.join(export_folder, "files")
                
                if os.path.exists(files_source):
                    shutil.copytree(files_source, files_dest)
                    progress.setValue(80)
            
            # Создаем README файл
            progress.setValue(90)
            progress.setLabelText("Создание описания...")
            
            readme_path = os.path.join(export_folder, "README.txt")
            stats = self.db_manager.get_documents_statistics()
            
            readme_text = f"""
ЭКСПОРТ БАЗЫ ДАННЫХ
===================

Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}

СОДЕРЖИМОЕ:
-----------
- {db_filename} - файл базы данных
- files/ - папка с документами (если включена)

СТАТИСТИКА:
-----------
Всего документов: {stats.get('total_documents', 0)}

ПО СТАТУСАМ:
{chr(10).join([f'  • {name}: {count}' for name, count in stats.get('by_status', {}).items()])}

ПО ТИПАМ:
{chr(10).join([f'  • {name}: {count}' for name, count in stats.get('by_type', {}).items()])}

ИНСТРУКЦИЯ ПО ВОССТАНОВЛЕНИЮ:
-----------------------------
1. Скопируйте папку экспорта в нужное место
2. Откройте приложение
3. Используйте кнопку "🔄 База данных" для переключения
4. Выберите файл {db_filename} из экспортированной папки

Или используйте функцию создания новой БД и укажите путь к этому файлу.
"""
            
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write(readme_text)
            
            progress.setValue(100)
            progress.close()
            
            # Сообщение об успехе
            QMessageBox.information(
                self,
                "✅ Экспорт завершен",
                f"База данных успешно экспортирована:\n\n"
                f"{export_folder}\n\n"
                f"Экспортировано документов: {stats.get('total_documents', 0)}"
            )
            
            self.accept()
            
            # Открыть папку
            reply = QMessageBox.question(
                self,
                "📂 Открыть папку?",
                "Открыть папку с экспортом?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.open_folder(export_folder)
        
        except Exception as e:
            if 'progress' in locals():
                progress.close()
            
            QMessageBox.critical(
                self,
                "❌ Ошибка экспорта БД",
                f"Не удалось экспортировать базу данных:\n\n{str(e)}"
            )
            print(f"❌ Ошибка экспорта БД: {e}")
            import traceback
            traceback.print_exc()
    
    def open_file(self, filepath: str):
        """Открыть файл в программе по умолчанию"""
        import platform
        import subprocess
        
        try:
            if platform.system() == 'Windows':
                os.startfile(filepath)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', filepath])
            else:
                subprocess.call(['xdg-open', filepath])
        except Exception as e:
            print(f"❌ Ошибка открытия файла: {e}")
    
    def open_folder(self, folder_path: str):
        """Открыть папку в проводнике"""
        import platform
        import subprocess
        
        try:
            if platform.system() == 'Windows':
                os.startfile(folder_path)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', folder_path])
            else:
                subprocess.call(['xdg-open', folder_path])
        except Exception as e:
            print(f"❌ Ошибка открытия папки: {e}")
    def set_documents_for_export(self, documents: list, mode: str = 'all_from_table'):
        """
        Установить список документов для экспорта
        
        Args:
            documents: Список документов (полные данные)
            mode: Режим экспорта ('all_from_table' или 'selected')
        """
        self.preset_documents = documents
        self.preset_mode = mode
        
        # Автоматически переключаемся на соответствующий режим
        if hasattr(self, 'preset_docs_radio'):
            self.preset_docs_radio.setChecked(True)
            self.on_period_mode_changed()
        
        print(f"✅ Установлено {len(documents)} документов для экспорта (режим: {mode})")


class ExportManager:
    """Менеджер экспорта документов в Excel"""
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def export_documents(self, 
                        documents: List[Dict], 
                        fields: List[tuple],
                        parent_widget=None) -> Optional[str]:
        """Экспортировать документы в Excel"""
        try:
            if not documents:
                return None
            
            # Генерируем имя файла
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"Реестр_документов_{timestamp}.xlsx"
            
            # Определяем папку сохранения
            db_folder = self.db_manager.db_folder
            default_path = os.path.join(db_folder, default_filename)
            
            # Диалог выбора места сохранения
            filepath, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "Сохранить реестр как",
                default_path,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not filepath:
                return None
            
            # Показываем прогресс
            progress = QProgressDialog(
                "Создание реестра...",
                "Отмена",
                0, 100,
                parent_widget
            )
            progress.setWindowModality(Qt.WindowModal)
            progress.setWindowTitle("📤 Экспорт")
            progress.show()
            
            # Создаем Excel файл
            progress.setValue(10)
            wb = Workbook()
            ws = wb.active
            ws.title = "Реестр документов"
            
            # Настройка стилей
            progress.setValue(20)
            self._setup_styles(ws)
            
            # Добавляем заголовок
            progress.setValue(30)
            self._add_header(ws, len(fields))
            
            # Добавляем названия колонок
            progress.setValue(40)
            self._add_column_headers(ws, fields)
            
            # Добавляем данные документов
            progress.setValue(50)
            total_docs = len(documents)
            for idx, doc in enumerate(documents):
                self._add_document_row(ws, doc, fields, idx + 3)
                
                progress_value = 50 + int((idx / total_docs) * 40)
                progress.setValue(progress_value)
                
                if progress.wasCanceled():
                    return None
            
            # Автоподбор ширины колонок
            progress.setValue(90)
            self._adjust_column_widths(ws, fields)
            
            # Сохраняем файл
            progress.setValue(95)
            wb.save(filepath)
            
            progress.setValue(100)
            progress.close()
            
            return filepath
            
        except Exception as e:
            if 'progress' in locals():
                progress.close()
            
            QMessageBox.critical(
                parent_widget,
                "❌ Ошибка экспорта",
                f"Не удалось создать реестр:\n\n{str(e)}"
            )
            print(f"❌ Ошибка экспорта: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _setup_styles(self, ws):
        """Настройка базовых стилей листа"""
        ws.freeze_panes = "A3"
    
    def _add_header(self, ws, num_columns: int):
        """Добавить заголовок реестра"""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
        
        header_cell = ws.cell(row=1, column=1)
        header_cell.value = "Реестр правовых актов администрации Сосновского муниципального района"
        
        header_cell.font = Font(
            name='Arial',
            size=14,
            bold=True,
            color='FFFFFF'
        )
        header_cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        header_cell.fill = PatternFill(
            start_color='366092',
            end_color='366092',
            fill_type='solid'
        )
        
        ws.row_dimensions[1].height = 40
    
    def _add_column_headers(self, ws, fields: List[tuple]):
        """Добавить названия колонок"""
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for col_idx, (field_key, field_name) in enumerate(fields, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = field_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[2].height = 30
    
    def _add_document_row(self, ws, document: Dict, fields: List[tuple], row_num: int):
        """Добавить строку с данными документа"""
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        if row_num % 2 == 0:
            cell_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        else:
            cell_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        for col_idx, (field_key, field_name) in enumerate(fields, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            
            value = self._get_field_value(document, field_key)
            cell.value = value
            
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.fill = cell_fill
            cell.border = thin_border
        
        ws.row_dimensions[row_num].height = 25
    
    def _get_field_value(self, document: Dict, field_key: str) -> str:
        """Получить значение поля документа"""
        if field_key == 'signers':
            signers = document.get('signers', [])
            if signers:
                return ', '.join([s.get('name', '') for s in signers])
            return '-'
        
        if field_key == 'approvers':
            approvers = document.get('approvers', [])
            if approvers:
                return ', '.join([a.get('name', '') for a in approvers])
            return '-'
        
        if field_key in ['reg_date', 'published_date', 'control_date']:
            date_value = document.get(field_key, '')
            if date_value:
                try:
                    date_obj = datetime.strptime(date_value, "%Y-%m-%d")
                    return date_obj.strftime("%d.%m.%Y")
                except:
                    return date_value
            return '-'
        
        value = document.get(field_key, '')
        
        if isinstance(value, bool):
            return 'Да' if value else 'Нет'
        
        if value is None or value == '':
            return '-'
        
        return str(value)
    
    def _adjust_column_widths(self, ws, fields: List[tuple]):
        """Автоматическая подстройка ширины колонок"""
        min_width = 10
        max_width = 50
        
        for col_idx, (field_key, field_name) in enumerate(fields, start=1):
            column_letter = get_column_letter(col_idx)
            
            if field_key in ['reg_number', 'number']:
                width = 15
            elif field_key in ['reg_date', 'published_date', 'control_date']:
                width = 12
            elif field_key == 'title':
                width = 40
            elif field_key in ['signers', 'approvers', 'executor_name', 'responsible_executor_name']:
                width = 25
            elif field_key in ['status_name', 'type_name', 'document_kind_name', 'signing_type_name', 'theme_name']:
                width = 20
            else:
                width = 15
            
            width = max(min_width, min(width, max_width))
            ws.column_dimensions[column_letter].width = width


def show_export_dialog(db_manager, parent=None) -> Optional[str]:
    """
    Показать диалог экспорта
    
    Args:
        db_manager: Менеджер базы данных
        parent: Родительский виджет
    
    Returns:
        str: Путь к созданному файлу или None
    """
    dialog = ExportDialog(db_manager, parent)
    dialog.exec_()
    return None
